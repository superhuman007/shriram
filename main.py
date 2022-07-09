import openpyxl
path="F:/HEMA/STFC/27Jun2022/pdf24_merged_daily.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row = 1, column = 1)
print(cell_obj.value)
res=len(wb_obj.sheetnames)
print(res)
if res<6 :
    print("2 executives in one sheet");