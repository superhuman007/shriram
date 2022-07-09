import openpyxl

print("PROCESSING DAILY FILE")
print("======================")
daily_mstr_file="F:/HEMA/STFC/27Jun2022/pdf24_merged_daily.xlsx"
mwb= openpyxl.load_workbook(daily_mstr_file)
no_of_sheets=len(mwb.sheetnames)
print('Number of Sheets in file: ',no_of_sheets)
if no_of_sheets<6 :
         print("2 executives in one sheet")
sheets=mwb.sheetnames
#first_sheet=mwb.get_sheet_names()[0]
print(sheets[0])
sheet=mwb[sheets[0]]
for c in sheet['B']:
    print(c.value)

#worksheet=mwb.
#for i in sheets:
    #print(i)
    #sheet=mwb["Table 1"]
    # for c in sheet['B']:
    #     cells=(c.value)
    #     j=0
    #     for k in cells:
    #         j=j+1
    #     print(j)

   # ws=mwb[sheets[0]]




# sheet_obj = wb1_obj.active
#cell_obj = sheet_obj.cell(row = 1, column = 1)
#print(cell_obj.value)
# res=len(wb1_obj.sheetnames)
# print(res)

# ws1=wb1_obj[0]
# print(mwb.sheetnames)
# dailyfile="F:/HEMA/STFC/27Jun2022/DRTL LCC.xlsx"
# wb2_obj=openpyxl.load_workbook(dailyfile)
# ws2=wb2_obj.active
# ws1_mr=ws1.max_row
# ws1_mc=ws1.max_column
# print(ws1_mc,ws1_mr)
